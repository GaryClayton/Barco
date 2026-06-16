from django.shortcuts import render

# Create your views here.
import os
from openpyxl import Workbook

from companys.models import Company, CompanyContact, ClubContact, CRM
from events.models import Events, Contribution
from charity.models import Charity, CharityContact, CharityDonation
from pages.models import Page


def write_excel_document(request):
    fpath = os.path.join('barc_root/uploads/', 'full_database_output.xlsx')

    if os.path.exists(fpath):
        try:
            os.remove(fpath)
        except:
            return render(request, 'reports/doc_failure.html', {'Page_list': Page.objects.all()})

    workbook = Workbook()

    sheet1 = workbook.active  # The default sheet
    sheet1.title = "Companies"
    sheet1.append(['ID', 'Name', 'Sector', 'Subsector', 'Address', 'Phone', 'email', 'web'])
    temp = Company.objects.all()
    for c in temp:
        sheet1.append([c.id, c.name, c.sector, c.subsector, c.address, c.phone, c.email, c.web])

    sheet2 = workbook.create_sheet(title = "Company Contacts")
    sheet2.append(['ID', 'Name', 'Position', 'Mobile', 'Phone', 'email', 'Company'])
    temp = CompanyContact.objects.all()
    for d in temp:
        sheet2.append([d.id, d.name, d.position, d.mobile, d.phone, d.email, str(d.company)])

    sheet3 = workbook.create_sheet(title="Club POC")
    sheet3.append(['ID', 'Name', 'Company'])
    temp = ClubContact.objects.all()
    for e in temp:
        sheet3.append([e.id, e.name, str(e.company)])

    sheet4 = workbook.create_sheet(title="CRM Entries")
    sheet4.append(['ID', 'UserName', 'Created', 'Company', 'Comment'])
    temp = CRM.objects.all()
    for d in temp:
        sheet4.append([d.id, str(d.username), str(d.created), str(d.company), d.comment])

    sheet5 = workbook.create_sheet(title="Events")
    sheet5.append(['ID', 'Name', 'Year', 'Date'])
    temp = Events.objects.all()
    for d in temp:
        sheet5.append([d.id, d.name, d.year, str(d.date)])

    sheet6 = workbook.create_sheet(title="Contributions")
    sheet6.append(['ID', 'Support', 'Value', 'Event', 'Company'])
    temp = Contribution.objects.all()
    for d in temp:
        sheet6.append([d.id, d.support, d.value, str(d.event), str(d.company)])

    sheet7 = workbook.create_sheet(title="Charities")
    sheet7.append(['ID', 'name', 'Sector', 'Overview', 'Web', 'Address', 'Phone'])
    temp = Charity.objects.all()
    for d in temp:
        sheet7.append([d.id, d.name, d.sector, d.overview, d.web, d.address, d.phone])

    sheet8 = workbook.create_sheet(title="Charity Contacts")
    sheet8.append(['ID', 'name', 'Title', 'Phone', 'Mobile', 'email', 'comment', 'Charity'])
    temp = CharityContact.objects.all()
    for d in temp:
        sheet8.append([d.id, d.name, d.title, d.phone, d.mobile, d.email, d.comment, str(d.charity)])

    sheet9 = workbook.create_sheet(title="Charity Donations")
    sheet9.append(['ID', 'name', 'Sector', 'Overview', 'Web', 'Address', 'Phone'])
    temp = CharityDonation.objects.all()
    for d in temp:
        sheet9.append([d.id, str(d.date), d.value, d.comment, str(d.charity)])


    workbook.save(fpath)
    return
